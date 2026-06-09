"""Generate Excel comparison: current 9-dashboard layout vs reshuffled 7-category layout."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "dashboard_comparison.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
MOVED_FILL = PatternFill("solid", fgColor="FFF2CC")
SAME_FILL = PatternFill("solid", fgColor="E2EFDA")
SPLIT_FILL = PatternFill("solid", fgColor="FCE4D6")

CURRENT_DASHBOARDS = [
    (1, "Request & Traffic Metrics", "ai-telemetry-executive"),
    (2, "Traffic & Request Analytics", "ai-telemetry-traffic"),
    (3, "Latency & Performance Metrics", "ai-telemetry-latency"),
    (4, "Cost & Usage Metrics", "ai-telemetry-cost"),
    (5, "Model Quality Metrics", "ai-telemetry-quality"),
    (6, "Safety & Security Metrics", "ai-telemetry-safety"),
    (7, "Infrastructure Metrics", "ai-telemetry-infra"),
    (8, "Token & Context Metrics", "ai-telemetry-tokens"),
    (9, "User-Level Observability", "ai-telemetry-users"),
]

RESHUFFLED_DASHBOARDS = [
    (1, "Infrastructure observability", "ai-telemetry-infra"),
    (2, "Network observability", "ai-telemetry-latency"),
    (3, "AI observability", "ai-telemetry-quality"),
    (4, "Data observability", "ai-telemetry-executive"),
    (5, "User observability", "ai-telemetry-users"),
    (6, "Cost & Usage Observability", "ai-telemetry-cost"),
    (7, "Safety and security", "ai-telemetry-safety"),
]

# metric, panel_type, top_card, cur_num, cur_name, new_num, new_name, new_section, top_after, change
METRICS: list[tuple] = [
    # ── D1 Request & Traffic ──
    ("Total requests", "Stat Card", "Yes", 1, "Request & Traffic Metrics", 4, "Data observability", "Top row", "Yes", "Moved"),
    ("Active users", "Stat Card", "Yes", 1, "Request & Traffic Metrics", 5, "User observability", "Top row", "Yes", "Moved"),
    ("Active sessions", "Stat Card", "Yes", 1, "Request & Traffic Metrics", 5, "User observability", "Top row", "Yes", "Moved"),
    ("RPM", "Timeseries", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Traffic", "No", "Moved"),
    ("RPM by Model", "Timeseries", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Traffic", "No", "Moved"),
    ("RPM by Model (current)", "Bar Gauge", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Traffic", "No", "Moved"),
    ("RPM Share by Model", "Pie Chart", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Traffic", "No", "Moved"),
    ("Error RPM by Model", "Timeseries", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Errors", "No", "Moved"),
    ("Error Rate by Model", "Timeseries", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Errors", "No", "Moved"),
    ("Model Provider Distribution", "Pie Chart", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Model mix", "No", "Moved"),
    ("Model Distribution (last 1h)", "Bar Gauge", "No", 1, "Request & Traffic Metrics", 4, "Data observability", "Model mix", "No", "Moved"),

    # ── D2 Traffic Analytics ──
    ("Requests / min by Model", "Timeseries", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Traffic Volume", "No", "Moved"),
    ("Requests / min by Department", "Timeseries", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Traffic Volume", "No", "Moved"),
    ("Requests / min by Operation", "Timeseries", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Traffic Volume", "No", "Moved"),
    ("Model Distribution (last 1h)", "Pie Chart", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Traffic Volume", "No", "Merged"),
    ("Model Provider Distribution", "Pie Chart", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Traffic Volume", "No", "Merged"),
    ("Error Rate by Type", "Timeseries", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Errors & Retries", "No", "Moved"),
    ("Errors by HTTP Status Code (last 1h)", "Bar Chart", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Errors & Retries", "No", "Moved"),
    ("Error Category Mix", "Pie Chart", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Errors & Retries", "No", "Moved"),
    ("SLA Breach Rate by Department", "Timeseries", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "SLA Compliance", "No", "Moved"),
    ("Total Requests by Routing Reason (last 1h)", "Bar Chart", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "SLA Compliance", "No", "Moved"),
    ("Live Telemetry Events", "Logs", "No", 2, "Traffic & Request Analytics", 4, "Data observability", "Live Request Log", "No", "Moved"),

    # ── D3 Latency ──
    ("End-to-end response latency", "Timeseries", "No", 3, "Latency & Performance Metrics", 2, "Network observability", "Latency", "No", "Moved"),
    ("Request Latency — p50 / p95 / p99", "Timeseries", "No", 3, "Latency & Performance Metrics", 2, "Network observability", "Latency", "No", "Moved"),
    ("Model Specific Latency", "Timeseries", "No", 3, "Latency & Performance Metrics", 3, "AI observability", "Model latency", "No", "Split"),
    ("First token latency (Model Based)", "Timeseries", "No", 3, "Latency & Performance Metrics", 3, "AI observability", "Model latency", "No", "Split"),
    ("Queue delays", "Timeseries", "No", 3, "Latency & Performance Metrics", 2, "Network observability", "Queue", "No", "Moved"),

    # ── D4 Cost ──
    ("Cost per request", "Timeseries", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Cost", "No", "Moved"),
    ("Cost per user/session", "Timeseries", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Cost", "No", "Moved"),
    ("Daily/monthly spend", "Timeseries", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Cost", "No", "Moved"),
    ("Cost by department", "Pie Chart", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Cost breakdown", "No", "Moved"),
    ("Model-wise cost breakdown", "Bar Chart", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Cost breakdown", "No", "Moved"),
    ("Cache hit savings", "Timeseries", "No", 4, "Cost & Usage Metrics", 6, "Cost & Usage Observability", "Savings", "No", "Moved"),

    # ── D5 Model Quality ──
    ("Hallucination rate", "Stat Card", "Yes", 5, "Model Quality Metrics", 3, "AI observability", "Top row", "Yes", "Moved"),
    ("Factual accuracy", "Stat Card", "Yes", 5, "Model Quality Metrics", 3, "AI observability", "Top row", "Yes", "Moved"),
    ("Relevance score", "Stat Card", "Yes", 5, "Model Quality Metrics", 3, "AI observability", "Top row", "Yes", "Moved"),
    ("Groundedness score", "Stat Card", "Yes", 5, "Model Quality Metrics", 3, "AI observability", "Top row", "Yes", "Moved"),
    ("Hallucination rate over time", "Timeseries", "No", 5, "Model Quality Metrics", 3, "AI observability", "Quality trends", "No", "Moved"),
    ("Factual accuracy", "Timeseries", "No", 5, "Model Quality Metrics", 3, "AI observability", "Quality trends", "No", "Moved"),
    ("Relevance score", "Timeseries", "No", 5, "Model Quality Metrics", 3, "AI observability", "Quality trends", "No", "Moved"),
    ("Groundedness score", "Timeseries", "No", 5, "Model Quality Metrics", 3, "AI observability", "Quality trends", "No", "Moved"),
    ("Evaluation Coverage", "Stat Card", "No", 5, "Model Quality Metrics", 3, "AI observability", "Evaluation Ops", "No", "Moved"),
    ("Evaluator Errors (24h)", "Stat Card", "No", 5, "Model Quality Metrics", 3, "AI observability", "Evaluation Ops", "No", "Moved"),
    ("Low-Quality Responses (hallucination flagged)", "Logs", "No", 5, "Model Quality Metrics", 3, "AI observability", "Evaluation Ops", "No", "Moved"),

    # ── D6 Safety ──
    ("Toxicity score", "Stat Card", "Yes", 6, "Safety & Security Metrics", 7, "Safety and security", "Top row", "Yes", "Renamed"),
    ("PII detection rate", "Stat Card", "Yes", 6, "Safety & Security Metrics", 7, "Safety and security", "Top row", "Yes", "Renamed"),
    ("Prompt injection attempts", "Stat Card", "Yes", 6, "Safety & Security Metrics", 7, "Safety and security", "Top row", "Yes", "Renamed"),
    ("Jailbreak attempts", "Stat Card", "Yes", 6, "Safety & Security Metrics", 7, "Safety and security", "Top row", "Yes", "Renamed"),
    ("Compliance violations", "Stat Card", "Yes", 6, "Safety & Security Metrics", 7, "Safety and security", "Top row", "Yes", "Renamed"),
    ("Toxicity score", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Safety Trends", "No", "Renamed"),
    ("PII detection rate", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Safety Trends", "No", "Renamed"),
    ("Prompt injection attempts / min", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Safety Trends", "No", "Renamed"),
    ("Jailbreak attempts / min", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Safety Trends", "No", "Renamed"),
    ("Compliance violations / min", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Safety Trends", "No", "Renamed"),
    ("PII Events Today", "Stat Card", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("PHI Requests Today", "Stat Card", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("PII Requests Today", "Stat Card", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("Unique Prompt Hashes (24h)", "Stat Card", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("Data Classification Distribution", "Pie Chart", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("PII Events by Department", "Timeseries", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("PHI + PII Volume by Department (last 1h)", "Bar Chart", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "PII & Data Classification", "No", "Renamed"),
    ("Prompt Log Events (PII-scrubbed)", "Logs", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "Prompt Audit Log", "No", "Renamed"),
    ("PHI / PII Requests with Trace Links", "Table", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "High-Risk Request Table", "No", "Renamed"),
    ("Safety incidents (injection, jailbreak, compliance)", "Logs", "No", 6, "Safety & Security Metrics", 7, "Safety and security", "High-Risk Request Table", "No", "Renamed"),

    # ── D7 Infrastructure ──
    ("CPU utilization", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("Model throughput", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("OOM failures", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("Pod/container health", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("Auto-scaling events", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("API error rate", "Stat Card", "Yes", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Top row", "Yes", "Moved"),
    ("CPU utilization", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("Model throughput", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("OOM failures", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("Pod/container health", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("Auto-scaling events", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("API error rate", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Infrastructure Trends", "No", "Moved"),
    ("HPA Current vs Desired Replicas", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "HPA Scaling", "No", "Moved"),
    ("Pod Restart Rate", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "HPA Scaling", "No", "Moved"),
    ("Container Memory RSS by Pod", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Container Resources", "No", "Moved"),
    ("Container CPU Usage by Pod", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Container Resources", "No", "Moved"),
    ("Node Memory Available Over Time", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Container Resources", "No", "Moved"),
    ("Node CPU Usage (user mode)", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Container Resources", "No", "Moved"),
    ("Batch Duration Heatmap", "Heatmap", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Runner Self-Observability", "No", "Moved"),
    ("Kafka Queue Depth", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Runner Self-Observability", "No", "Moved"),
    ("Runner Event Publish Rate", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Runner Self-Observability", "No", "Moved"),
    ("Batch Duration p99", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "Runner Self-Observability", "No", "Moved"),
    ("Collector Exporter Queue Size", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "OTel Collector", "No", "Moved"),
    ("Collector Export Throughput & Failures", "Timeseries", "No", 7, "Infrastructure Metrics", 1, "Infrastructure observability", "OTel Collector", "No", "Moved"),

    # ── D8 Token & Context ──
    ("Output tokens", "Stat Card", "Yes", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Top row", "Yes", "Moved"),
    ("Avg tokens / request", "Stat Card", "Yes", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Top row", "Yes", "Moved"),
    ("Context fill", "Stat Card", "Yes", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Top row", "Yes", "Moved"),
    ("Errors (5m)", "Stat Card", "Yes", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Top row", "Yes", "Moved"),
    ("Output Token Count", "Timeseries", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Output throughput", "No", "Moved"),
    ("Output tokens by model (now)", "Bar Gauge", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Output throughput", "No", "Moved"),
    ("Token type mix", "Pie Chart", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Request composition", "No", "Moved"),
    ("Total tokens per request", "Timeseries", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Request composition", "No", "Moved"),
    ("Prompt size by model", "Bar Gauge", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Request composition", "No", "Moved"),
    ("Context Window Utilization (%)", "Gauge", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Context window", "No", "Moved"),
    ("Context fill by model", "Bar Gauge", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Context window", "No", "Moved"),
    ("Prompt size trend", "Timeseries", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Context window", "No", "Moved"),
    ("Live token generation rate", "Timeseries", "No", 8, "Token & Context Metrics", 3, "AI observability", "Streaming performance", "No", "Split"),
    ("Streaming response latency", "Timeseries", "No", 8, "Token & Context Metrics", 3, "AI observability", "Streaming performance", "No", "Split"),
    ("Streaming tokens/sec", "Bar Gauge", "No", 8, "Token & Context Metrics", 3, "AI observability", "Streaming throughput", "No", "Split"),
    ("Avg stream tokens/s", "Stat Card", "No", 8, "Token & Context Metrics", 3, "AI observability", "Streaming throughput", "No", "Split"),
    ("Errors by type (5m)", "Bar Chart", "No", 8, "Token & Context Metrics", 6, "Cost & Usage Observability", "Streaming throughput & errors", "No", "Moved"),

    # ── D9 User ──
    ("Logins (24h)", "Stat Card", "Yes", 9, "User-Level Observability", 5, "User observability", "Top row", "Yes", "Moved"),
    ("Active users (24h)", "Stat Card", "Yes", 9, "User-Level Observability", 5, "User observability", "Top row", "Yes", "Moved"),
    ("Monthly active users (30d)", "Stat Card", "Yes", 9, "User-Level Observability", 5, "User observability", "Top row", "Yes", "Moved"),
    ("LLM usage spike (15m vs prev 15m)", "Stat Card", "Yes", 9, "User-Level Observability", 5, "User observability", "Top row", "Yes", "Moved"),
    ("Login track", "Timeseries", "No", 9, "User-Level Observability", 5, "User observability", "Login & user growth", "No", "Moved"),
    ("Daily active users (24h)", "Stat Card", "No", 9, "User-Level Observability", 5, "User observability", "Login & user growth", "No", "Moved"),
    ("Active users by department (24h)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Login & user growth", "No", "Moved"),
    ("Top 10 users — tokens (24h)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Token consumption by user", "No", "Moved"),
    ("Top 10 users — token rate (5m)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Token consumption by user", "No", "Moved"),
    ("Top 10 users — session time (6h)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Session-level usage", "No", "Moved"),
    ("Top users by tokens (6h)", "Table", "No", 9, "User-Level Observability", 5, "User observability", "Session-level usage", "No", "Moved"),
    ("Session time by user (top 10, 5m)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Session-level usage", "No", "Moved"),
    ("Token volume — spike detector (5m buckets)", "Timeseries", "No", 9, "User-Level Observability", 5, "User observability", "Usage spikes", "No", "Moved"),
    ("Top 10 users — spike ratio (15m vs prev 15m)", "Bar Chart", "No", 9, "User-Level Observability", 5, "User observability", "Usage spikes", "No", "Moved"),
    ("Recent login events", "Logs", "No", 9, "User-Level Observability", 5, "User observability", "Usage spikes", "No", "Moved"),
]


def _style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 45)


def sheet_summary(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Dashboard Summary"
    headers = [
        "Version", "#", "Dashboard Name", "UID", "In Nav?",
        "Panel Count", "Top Stat Cards", "Notes",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    counts = {}
    top_cards = {}
    for m in METRICS:
        key = (m[3], m[4])
        counts[key] = counts.get(key, 0) + 1
        if m[2] == "Yes":
            top_cards[key] = top_cards.get(key, 0) + 1

    for num, name, uid in CURRENT_DASHBOARDS:
        key = (num, name)
        in_nav = "Yes"
        notes = ""
        if num in (2, 8):
            notes = "Dropped from nav in reshuffled version (metrics absorbed elsewhere)"
        ws.append([
            "Current (9 dashboards)", num, name, uid, in_nav,
            counts.get(key, 0), top_cards.get(key, 0), notes,
        ])

    ws.append([])
    reshuffle_counts: dict[tuple, int] = {}
    reshuffle_tops: dict[tuple, int] = {}
    for m in METRICS:
        key = (m[5], m[6])
        reshuffle_counts[key] = reshuffle_counts.get(key, 0) + 1
        if m[8] == "Yes":
            reshuffle_tops[key] = reshuffle_tops.get(key, 0) + 1

    for num, name, uid in RESHUFFLED_DASHBOARDS:
        key = (num, name)
        notes = ""
        if num == 2:
            notes = "No top stat cards — charts only (per Excel)"
        if num == 4:
            notes = "Absorbs #1 Request & Traffic + #2 Traffic Analytics"
        if num == 5:
            notes = "Absorbs Active users/sessions from #1 + all of #9"
        if num == 6:
            notes = "Absorbs #4 Cost + #8 Token & Context (except streaming → AI)"
        ws.append([
            "Reshuffled (7 dashboards)", num, name, uid, "Yes",
            reshuffle_counts.get(key, 0), reshuffle_tops.get(key, 0), notes,
        ])

    _autosize(ws)


def sheet_metrics(wb: Workbook) -> None:
    ws = wb.create_sheet("All Metrics Compare")
    headers = [
        "Metric / Panel", "Panel Type",
        "Current #", "Current Dashboard",
        "Top Card (Current)?",
        "Reshuffled #", "Reshuffled Dashboard",
        "Reshuffled Section", "Top Card (After)?",
        "Change Type",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    fill_map = {
        "Same": SAME_FILL,
        "Moved": MOVED_FILL,
        "Split": SPLIT_FILL,
        "Merged": MOVED_FILL,
        "Renamed": SAME_FILL,
    }

    for row_idx, m in enumerate(METRICS, start=2):
        ws.append([m[0], m[1], m[3], m[4], m[2], m[5], m[6], m[7], m[8], m[9]])
        fill = fill_map.get(m[9], MOVED_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def sheet_top_cards(wb: Workbook) -> None:
    ws = wb.create_sheet("Top Cards Compare")
    headers = [
        "Metric", "Current Dashboard", "Current Position",
        "Reshuffled Dashboard", "Reshuffled Position", "Change",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    tops = [m for m in METRICS if m[2] == "Yes" or m[8] == "Yes"]
    seen = set()
    for m in tops:
        key = (m[0], m[3], m[5])
        if key in seen:
            continue
        seen.add(key)
        cur_pos = "Top row" if m[2] == "Yes" else "Secondary row"
        new_pos = "Top row" if m[8] == "Yes" else ("Secondary row" if m[2] == "Yes" else "—")
        ws.append([m[0], f"#{m[3]} {m[4]}", cur_pos, f"#{m[5]} {m[6]}", new_pos, m[9]])

    _autosize(ws)


def sheet_nav(wb: Workbook) -> None:
    ws = wb.create_sheet("Nav Bar Compare")
    headers = ["Nav #", "Current Label", "Current UID", "Reshuffled Label", "Reshuffled UID", "Status"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    mapping = [
        (1, "Request & Traffic", "ai-telemetry-executive", "Infrastructure", "ai-telemetry-infra", "Reordered + renamed"),
        (2, "Traffic Analytics", "ai-telemetry-traffic", "Network", "ai-telemetry-latency", "Reordered + renamed"),
        (3, "Latency", "ai-telemetry-latency", "AI", "ai-telemetry-quality", "Reordered + renamed"),
        (4, "Cost & Usage", "ai-telemetry-cost", "Data", "ai-telemetry-executive", "Reordered + renamed"),
        (5, "Model Quality", "ai-telemetry-quality", "User", "ai-telemetry-users", "Reordered + renamed"),
        (6, "Safety & Security", "ai-telemetry-safety", "Cost & Usage", "ai-telemetry-cost", "Reordered + renamed"),
        (7, "Infrastructure", "ai-telemetry-infra", "Safety & Security", "ai-telemetry-safety", "Reordered + renamed"),
        (8, "Token & Context", "ai-telemetry-tokens", "—", "—", "Removed from nav (merged into Cost & AI)"),
        (9, "User Observability", "ai-telemetry-users", "—", "—", "Removed from nav (merged into User)"),
    ]
    for row in mapping:
        ws.append(row)
    _autosize(ws)


def sheet_by_reshuffled(wb: Workbook) -> None:
    ws = wb.create_sheet("By Reshuffled Dashboard")
    headers = [
        "Reshuffled #", "Reshuffled Dashboard", "Metric / Panel",
        "Panel Type", "Top Card?", "From Current Dashboard(s)", "Section",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for num, name, _uid in RESHUFFLED_DASHBOARDS:
        rows = [m for m in METRICS if m[5] == num]
        for m in rows:
            ws.append([
                num, name, m[0], m[1],
                "Yes" if m[8] == "Yes" else "No",
                f"#{m[3]} {m[4]}", m[7],
            ])
        if rows:
            ws.append([])

    _autosize(ws)


def main() -> None:
    wb = Workbook()
    sheet_summary(wb)
    sheet_metrics(wb)
    sheet_top_cards(wb)
    sheet_nav(wb)
    sheet_by_reshuffled(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Metrics compared: {len(METRICS)}")


if __name__ == "__main__":
    main()
